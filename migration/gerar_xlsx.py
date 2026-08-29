# -*- coding: utf-8 -*-
"""Planilha da Fase 2 a partir do inventario da Fase 1.

    python gerar_xlsx.py --src <dir-do-export> --out <arquivo.xlsx>

Mesma ferramenta e paleta de docs/novo-modelo/exportacoes/gerar_xlsx.py
(openpyxl), para a planilha de migracao nao destoar dos relatorios do painel.

A coluna DECISAO tem validacao por linha: so aceita os codigos validos daquele
caso. Sem isso a planilha volta da revisao com destino que nao existe.

CONTEM NOME DE PACIENTE. Grave fora do repositorio.
"""
import argparse
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from catalogo_novo import CARDS
from fase1_inventario import classificar, derivacao_card06, ler_pacientes

# Paleta de docs/novo-modelo/exportacoes/exp_common.py
PINE, PINE_L = "1F5F52", "E4EFEA"
BRICK, BRICK_L = "A3462F", "F6E9E4"
INK, INK2, INK3 = "15211D", "47554F", "77857E"
LINE, SAND = "E0E5E1", "9A7B1F"

BORDA = Border(bottom=Side(style="thin", color=LINE))
H_FILL = PatternFill("solid", fgColor=PINE)
H_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
T_FONT = Font(name="Calibri", size=16, bold=True, color=INK)
S_FONT = Font(name="Calibri", size=10, color=INK3)
CORPO = Font(name="Calibri", size=10, color=INK2)

# Os valores de `tipo` e `confianca` sao identificadores usados em comparacao;
# ficam ASCII. Acentuado e so o rotulo que chega ao usuario.
ROTULO_TIPO = {"direto": "direto", "derivacao": "derivação",
               "ambiguo": "ambíguo", "sem_regra": "sem regra"}
ROTULO_CONF = {"alta": "alta", "media": "média", "baixa": "baixa",
               "nenhuma": "nenhuma"}

COR_CONFIANCA = {
    "alta": (PINE_L, PINE),
    "media": ("FFF4D6", SAND),
    "baixa": ("F2F4F3", INK3),
    "nenhuma": (BRICK_L, BRICK),
}


def cabecalho(ws, titulo, sub):
    ws["A1"] = titulo
    ws["A1"].font = T_FONT
    ws["A2"] = sub
    ws["A2"].font = S_FONT
    ws.sheet_view.showGridLines = False


def largura(ws, linhas, limites):
    for i, teto in enumerate(limites, 1):
        m = 10
        for row in linhas:
            if i <= len(row) and row[i - 1] is not None:
                m = max(m, min(teto, len(str(row[i - 1])) + 3))
        ws.column_dimensions[get_column_letter(i)].width = m


def tabela(ws, linha0, headers, linhas):
    for c, h in enumerate(headers, 1):
        cel = ws.cell(row=linha0, column=c, value=h)
        cel.fill, cel.font = H_FILL, H_FONT
        cel.alignment = Alignment(vertical="center")
    for r, row in enumerate(linhas, linha0 + 1):
        for c, v in enumerate(row, 1):
            cel = ws.cell(row=r, column=c, value=v)
            cel.font, cel.border = CORPO, BORDA
            cel.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws.cell(row=linha0 + 1, column=1)


def aba_resumo(wb, linhas, competencia):
    ws = wb.active
    ws.title = "Resumo"
    cabecalho(ws, "Migração — inventário da Fase 1", "Dump de %s · %d eventos" % (competencia, len(linhas)))
    por_tipo = Counter(l["tipo"] for l in linhas)
    dados = [[ROTULO_TIPO[t], por_tipo.get(t, 0), "%.1f%%" % (100 * por_tipo.get(t, 0) / len(linhas)), d]
             for t, d in [
                 ("direto", "Equivalência 1:1 — o loader resolve sozinho"),
                 ("derivacao", "Não vira fato no modelo novo; alimenta outra estrutura"),
                 ("ambiguo", "O modelo novo pede distinção que o dado velho não tem"),
                 ("sem_regra", "Nenhuma regra do de-para cobre"),
             ] if por_tipo.get(t)]
    tabela(ws, 4, ["Classe", "Eventos", "%", "Significado"], dados)

    amb = [l for l in linhas if l["tipo"] in ("ambiguo", "sem_regra")]
    conf = Counter(l["confianca"] for l in amb)
    ws.cell(row=4 + len(dados) + 3, column=1, value="Sugestão automática nos %d ambíguos" % len(amb)).font = Font(
        name="Calibri", size=11, bold=True, color=INK)
    tabela(ws, 4 + len(dados) + 4, ["Confiança", "Linhas", "Como foi obtida"], [
        ["alta", conf.get("alta", 0), "o próprio registro rotula o teor"],
        ["média", conf.get("media", 0), "palavra-chave forte no texto livre"],
        ["baixa", conf.get("baixa", 0), "evidência indireta (situação do paciente, evento vizinho)"],
        ["nenhuma", conf.get("nenhuma", 0), "nada no dado sustenta palpite — decisão 100% humana"],
    ])
    largura(ws, dados, [16, 10, 10, 62])


def aba_depara(wb, linhas):
    ws = wb.create_sheet("De-para")
    cabecalho(ws, "De-para por origem", "Contagem real sobre o dump")
    agrupado = defaultdict(lambda: [0, None])
    for l in linhas:
        k = (l["prefixo_antigo"], l["subindicador_antigo"] or l["indicador_antigo"])
        agrupado[k][0] += 1
        agrupado[k][1] = l
    dados = []
    for (pref, rotulo), (n, ex) in sorted(agrupado.items()):
        alvo = ("%s %s" % (ex["destino"], ex["destino_nome"])).strip() or ex["opcoes"] or "?"
        dados.append([pref, rotulo, n, ROTULO_TIPO[ex["tipo"]], alvo, ex["nota"]])
    tabela(ws, 4, ["Prefixo", "Categoria antiga", "Eventos", "Classe", "Destino", "Nota"], dados)
    largura(ws, dados, [12, 40, 10, 12, 44, 70])


def aba_ambiguos(wb, linhas):
    amb = [l for l in linhas if l["tipo"] in ("ambiguo", "sem_regra")]
    ws = wb.create_sheet("Ambíguos (Fase 2)")
    cabecalho(ws, "Ambíguos — revisão humana",
              "%d eventos. Preencha DECISAO; a sugestão é apoio, não decisão." % len(amb))
    headers = ["Evento", "Paciente", "Data", "Categoria antiga", "Observações",
               "Pista", "Sugestão", "Confiança", "Por quê", "Opções", "DECISAO"]
    dados = [[l["evento_id"][:12], l["paciente"], l["data"], l["subindicador_antigo"],
              l["observacoes_raw"], l["pista"],
              ("%s %s" % (l["sugestao"], l["sugestao_nome"])).strip(),
              ROTULO_CONF.get(l["confianca"], l["confianca"]),
              l["motivo_sugestao"], l["opcoes"], l["sugestao"]]
             for l in amb]
    tabela(ws, 4, headers, dados)

    col_conf, col_dec = 8, 11
    for r, l in enumerate(amb, 5):
        fundo, texto = COR_CONFIANCA.get(l["confianca"], ("FFFFFF", INK2))
        cel = ws.cell(row=r, column=col_conf)
        cel.fill = PatternFill("solid", fgColor=fundo)
        cel.font = Font(name="Calibri", size=10, bold=True, color=texto)

        # Validacao por linha: so os codigos validos deste caso.
        codigos = [o.split()[0] for o in l["opcoes"].split(" | ") if o.strip()]
        if not codigos and l["destino"]:
            codigos = [l["destino"]]
        if codigos:
            dv = DataValidation(type="list", formula1='"%s"' % ",".join(codigos),
                                allow_blank=True, showErrorMessage=True,
                                errorTitle="Destino inválido",
                                error="Use um dos códigos: " + ", ".join(codigos))
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=r, column=col_dec))
        d = ws.cell(row=r, column=col_dec)
        d.font = Font(name="Calibri", size=10, bold=True, color=INK)
        d.fill = PatternFill("solid", fgColor="FFFDF0")
    largura(ws, dados, [14, 30, 12, 30, 60, 46, 26, 12, 46, 40, 12])


def aba_card06(wb, card06):
    ws = wb.create_sheet("Card 06 (derivação)")
    cabecalho(ws, "Card 06 — AD/ID não vira fato",
              "%d eventos derivados em modalidade de episódio" % len(card06))
    dados = [[c["paciente"], c["data"], c["subindicador_antigo"], c["modalidade"], c["papel"]]
             for c in card06]
    tabela(ws, 4, ["Paciente", "Data", "Subindicador antigo", "Modalidade", "Papel"], dados)
    largura(ws, dados, [30, 12, 34, 14, 34])


def aba_catalogo(wb):
    ws = wb.create_sheet("Catálogo novo")
    cabecalho(ws, "Recategorização — catálogo", "Transcrito do PDF do novo modelo")
    dados = []
    for code, (nome, subs, nota) in CARDS.items():
        dados.append([code, nome, "", "", nota])
        for sub, rotulo in subs.items():
            dados.append(["", "", sub, rotulo, ""])
    tabela(ws, 4, ["Card", "Nome do card", "Código", "Subcategoria", "Nota"], dados)
    largura(ws, dados, [10, 34, 10, 46, 76])


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--competencia", default="28/08/2026")
    args = ap.parse_args()

    pacientes = ler_pacientes(args.src)
    linhas = classificar(pacientes)
    card06 = derivacao_card06(pacientes)

    wb = Workbook()
    aba_resumo(wb, linhas, args.competencia)
    aba_depara(wb, linhas)
    aba_ambiguos(wb, linhas)
    aba_card06(wb, card06)
    aba_catalogo(wb)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print("gravado: %s" % args.out)
    print("  abas: Resumo, De-para, Ambíguos (Fase 2), Card 06, Catálogo novo")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
