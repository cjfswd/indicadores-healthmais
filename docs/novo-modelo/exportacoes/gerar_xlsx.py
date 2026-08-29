# -*- coding: utf-8 -*-
from exp_common import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
thin = Side(style='thin', color=LINE)
BORDA = Border(bottom=thin)
H_FILL = PatternFill('solid', fgColor=PINE)
H_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
T_FONT = Font(name='Calibri', size=16, bold=True, color=INK)
S_FONT = Font(name='Calibri', size=10, color=INK3)
B_FONT = Font(name='Calibri', size=11, bold=True, color=INK)

def largura(ws, dados, cols):
    for i in range(cols):
        m = 10
        for row in dados:
            if i < len(row) and row[i] is not None:
                m = max(m, min(52, len(str(row[i])) + 3))
        ws.column_dimensions[get_column_letter(i + 1)].width = m

def cabecalho(ws, titulo, sub):
    ws['A1'] = titulo
    ws['A1'].font = T_FONT
    ws['A2'] = sub
    ws['A2'].font = S_FONT
    ws.sheet_view.showGridLines = False

def tabela(ws, linha0, headers, linhas, nome_tab=None):
    for c, h in enumerate(headers, 1):
        cel = ws.cell(row=linha0, column=c, value=h)
        cel.fill = H_FILL
        cel.font = H_FONT
        cel.alignment = Alignment(vertical='center')
    for r, row in enumerate(linhas, linha0 + 1):
        for c, v in enumerate(row, 1):
            cel = ws.cell(row=r, column=c)
            n = num(v) if isinstance(v, str) else v
            if isinstance(v, str) and is_money(v) and n is not None:
                cel.value = n
                cel.number_format = 'R$ #,##0.00'
            elif isinstance(v, str) and is_pct(v) and n is not None:
                cel.value = n / 100
                cel.number_format = '0.0%'
            elif isinstance(v, str) and n is not None and v.strip() == str(n):
                cel.value = n
            else:
                cel.value = v
            cel.font = Font(name='Calibri', size=10, color=INK2)
            cel.border = BORDA
    fim = linha0 + len(linhas)
    if nome_tab and linhas:
        ref = "A%d:%s%d" % (linha0, get_column_letter(len(headers)), fim)
        t = Table(displayName=nome_tab, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(t)
    ws.freeze_panes = ws.cell(row=linha0 + 1, column=1)
    largura(ws, [headers] + linhas, len(headers))
    return fim

ws = wb.active
ws.title = "Resumo"
cabecalho(ws, "Painel de Indicadores - competencia " + COMP,
          "Gerado em 28/08/2026 - recorte: todos os periodos, todas as operadoras, AD e ID")
linhas = []
for c in CARDS:
    st = c['stats']
    def g(i, campo):
        return st[i][campo] if len(st) > i else ''
    linhas.append([c['code'], c['nome'], g(0, 'v'), g(0, 'k'), g(1, 'v'), g(1, 'k'),
                   g(2, 'v'), g(2, 'k'), g(3, 'v'), g(3, 'k'),
                   (c['meta']['txt'] if c['meta'] else '-'),
                   ('Atingida' if c['meta'] and c['meta']['ok'] else ('Fora da meta' if c['meta'] else '-'))])
fim = tabela(ws, 4, ["Card", "Indicador", "Principal", "Rotulo", "Metrica 2", "Rotulo 2",
                     "Metrica 3", "Rotulo 3", "Metrica 4", "Rotulo 4", "Meta", "Situacao"],
             linhas, "Resumo")
for r in range(5, fim + 1):
    v = ws.cell(row=r, column=12).value
    if v == 'Fora da meta':
        ws.cell(row=r, column=12).font = Font(name='Calibri', size=10, bold=True, color=BRICK)
    elif v == 'Atingida':
        ws.cell(row=r, column=12).font = Font(name='Calibri', size=10, bold=True, color=PINE)

for c in CARDS:
    nome = (c['code'] + " " + c['nome'])[:31].replace('/', '-')
    w = wb.create_sheet(nome)
    cabecalho(w, c['code'] + " - " + c['nome'], c['nota'])
    tabela(w, 4, c['colunas'], c['linhas'], "T" + c['code'])

w = wb.create_sheet("Pivo mensal")
cabecalho(w, "Evolucao mensal por indicador e subindicador", "Contagem por mes de ocorrencia - card 07 em R$")
linha = 4
for c in CARDS:
    piv = c['pivot']
    cel = w.cell(row=linha, column=1, value=c['code'] + " - " + c['nome'])
    cel.font = B_FONT
    linha += 1
    for i, h in enumerate(piv['cols'], 1):
        x = w.cell(row=linha, column=i, value=h)
        x.fill = H_FILL
        x.font = H_FONT
    linha += 1
    for row in piv['rows']:
        for i, v in enumerate(row['cells'], 1):
            x = w.cell(row=linha, column=i)
            n = num(v)
            if is_money(v) and n is not None:
                x.value = n
                x.number_format = 'R$ #,##0.00'
            elif n is not None:
                x.value = n
            else:
                x.value = v
            x.font = Font(name='Calibri', size=10, bold=row['pai'], color=INK if row['pai'] else INK2)
            if row['pai']:
                x.fill = PatternFill('solid', fgColor=PINE_L)
            x.border = BORDA
        linha += 1
    linha += 1
w.column_dimensions['A'].width = 46
for i in range(2, 16):
    w.column_dimensions[get_column_letter(i)].width = 13
w.sheet_view.showGridLines = False

w = wb.create_sheet("Alertas")
cabecalho(w, "Pendencias e alertas em aberto", "Regras do documento de recategorizacao aplicadas ao periodo")
al = []
for c in CARDS:
    for a in c['alertas']:
        al.append([c['code'], c['nome'], a['registro'], a['quem'], a['texto']])
tabela(w, 4, ["Card", "Indicador", "Registro", "Paciente", "Pendencia"], al, "Alertas")
for r in range(5, 5 + len(al)):
    w.cell(row=r, column=5).font = Font(name='Calibri', size=10, color=BRICK)
w.column_dimensions['E'].width = 96

w = wb.create_sheet("Metas")
cabecalho(w, "Metas por indicador", "targetType, targetDirection e targetValue preservados do modelo atual")
tabela(w, 4, ["Card", "Indicador", "Meta", "Situacao"],
       [[c['code'], c['nome'], c['meta']['txt'] if c['meta'] else '-',
         'Atingida' if c['meta'] and c['meta']['ok'] else ('Fora da meta' if c['meta'] else 'Sem meta - card espelho')]
        for c in CARDS], "Metas")

w = wb.create_sheet("Retrocompatibilidade")
cabecalho(w, "O que vem do sistema atual", "Inventario por pagina: mantido, muda e novo")
rows = []
for p in D['paginas']:
    for k, v in p['itens']:
        rows.append([p['nome'], k, v])
tabela(w, 4, ["Pagina", "Situacao", "Descricao"], rows, "Retro")
w.column_dimensions['C'].width = 110

wb.save("Painel_Indicadores_08-2026.xlsx")
print("xlsx ok -", len(wb.sheetnames), "abas")
